# coding: utf-8
# Author: Jingcheng Qiu

"""
自动编辑excel，支持将DataFrame和图片插入excel
"""


import os
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter, coordinate_to_tuple


class Cell(object):
    def __init__(self, coords):
        """
        Parameters
        ----------
        coords: str or tuple
            单元格位置, 可以传入Excel单元格的格式, eg. 'A1'，也可传入单元格的坐标, eg. '(3, 1)'表示第3行第1列
        """
        if isinstance(coords, tuple):
            self.row = coords[0]
            self.col = coords[1]
            self.name = get_column_letter(coords[1]) + str(coords[0])
        elif isinstance(coords, str):
            self.name = coords.upper()
            self.row, self.col = coordinate_to_tuple(self.name)
        else:
            raise ValueError('coords type must be tuple or str')

    def shift(self, m, n):
        """
        输出单元格移动m行，n列后的单元格名称
        Parameters
        ----------
        m: int
            水平移动列数
        n: int
            垂直移动行数

        Returns
        -------
        name: str
            单元格名称
        """
        row = self.row + m
        col = self.col + n
        if row <= 0 or col <= 0:
            raise ValueError('row or column out of range after shift')
        name = get_column_letter(col) + str(row)

        return name

    def insert_df(self, df, sheet, header=False):
        """
        逐列将DataFrame的内容写入excel表格指定区域
        Parameters
        ----------
        df: pandas.DataFrame
            需要写入的dataframe
        sheet: object
            active excel sheet
        header: bool, default False
            是否要写入DataFrame的列名
        """
        offset = 0
        if header:
            for col, colname in enumerate(df.columns.tolist()):
                sheet[self.shift(0, col)] = colname
            offset = 1

        # 将dataframe按列写入
        for col, colname in enumerate(df.columns.tolist()):
            for row, value in enumerate(df[colname].tolist()):
                sheet[self.shift(offset + row, col)] = value

    def insert_image(self, image_path, sheet, img_size):
        """
        将图片插入到excel表格指定区域
        Parameters
        ----------
        image_path: str
            图片路径
        sheet: object
            active excel sheet
        img_size: tuple
            图片大小，占据表格的几行几列
        """
        # 计算将图片插入指定大小的区域时的像素(openpyxl插入图片时会被默认放大1.25倍，所以在插入时应该先缩小1.25倍)
        img = Image(image_path)
        img.drawing.width = int(img_size[1] * 80 / 1.25)  # excel一列为80像素
        img.drawing.height = int(img_size[0] * 23 / 1.25)  # excel一行为23像素
        sheet.add_image(img, self.name)


def insert_df(anchor, df, sheet, header=False):
    """
    将DataFrame的内容写入excel表格指定区域
    Parameters
    ----------
    anchor: str
        写入dataframe的顶点位置, eg.'A1'
    df: pandas.DataFrame
        需要写入的dataframe
    sheet: object
        active excel sheet
    header: bool, default False
        是否要写入DataFrame的列名
    """
    vertex = Cell(anchor)
    vertex.insert_df(df, sheet=sheet, header=header)


def batch_insert_image(anchor, images, sheet, img_size, interval=1, expand='vertical'):
    """
    批量插入图片到excel表格
    Parameters
    ----------
    anchor: str
        第一张图片的起始单元格位置, eg. 'A1'
    images: list or str
        图片路径，当为list时，直接对list中每个图片循环插入，当为文件夹路径时，将文件夹中的jpg、png格式的图片循环插入
    sheet: object
        active excel sheet
    img_size: tuple
        图片大小，占据表格的几行几列
    interval: int, default 1
        图片间的间隔距离，单位是行或列
    expand: str, default 'vertical', options ['vertical', 'horizontal']
        图片的排列方式，'vertical'为纵向排列, 'horizontal'为横向排列
    """
    vertex = Cell(anchor)

    if isinstance(images, list):
        image_list = images
    elif os.path.isdir(images):
        image_list = []
        for img in os.listdir(images):
            if img.split('.')[-1] in ['jpg', 'png']:
                image_list.append(os.path.join(images, img))
    else:
        raise ValueError('Wrong type images')

    for img in image_list:
        vertex.insert_image(img, sheet, img_size)
        if expand == 'vertical':
            anchor = vertex.shift(img_size[0] + interval, 0)
        elif expand == 'horizontal':
            anchor = vertex.shift(0, img_size[1] + interval)
        vertex = Cell(anchor)
    else:
        return True
